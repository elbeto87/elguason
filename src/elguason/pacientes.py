"""Lectura del Excel de pacientes desde el escritorio.

El Excel debe llamarse ``pacientes`` (``pacientes.xlsx``) y estar ubicado en el
escritorio del usuario (``~/Desktop`` o ``~/Escritorio``).

Se espera la siguiente estructura de columnas (con encabezado en la primera fila):

    nombre y apellido | cuit | numero de sesiones | honorarios por sesion | medio de pago | total

Donde ``total`` = ``numero de sesiones`` * ``honorarios por sesion``.
Si la columna ``total`` viene vacía se calcula automáticamente.
``medio de pago`` es el medio de pago a informar en la factura (ej: "Contado",
"Tarjeta de débito", "Transferencia", etc.).
"""
from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional

from loguru import logger
from openpyxl import load_workbook


NOMBRE_ARCHIVO_PACIENTES = "pacientes.xlsx"



@dataclass
class Paciente:
    """Representa una fila del Excel de pacientes."""
    nombre_y_apellido: str
    cuit: str
    numero_de_sesiones: int
    honorarios_por_sesion: int
    medio_de_pago: str = ""

    @property
    def total(self) -> int:
        """Total a facturar = sesiones * honorarios por sesión."""
        return self.numero_de_sesiones * self.honorarios_por_sesion


def _buscar_escritorio() -> Path:
    """Devuelve la ruta al escritorio del usuario.

    Contempla instalaciones en inglés (``Desktop``) y en español (``Escritorio``).
    """
    home = Path.home()
    for nombre in ("Desktop", "Escritorio"):
        candidato = home / nombre
        if candidato.is_dir():
            return candidato
    # Fallback a Desktop aunque no exista, para dar un mensaje de error claro
    return home / "Desktop"


def ruta_excel_pacientes(escritorio: Optional[Path] = None) -> Path:
    """Devuelve la ruta esperada del Excel de pacientes en el escritorio."""
    escritorio = escritorio or _buscar_escritorio()
    return escritorio / NOMBRE_ARCHIVO_PACIENTES


def _to_int(value, campo: str, fila: int) -> int:
    """Convierte un valor de celda a int validando que no sea vacío."""
    if value is None or (isinstance(value, str) and not value.strip()):
        raise ValueError(f"Falta '{campo}' en la fila {fila} del Excel de pacientes")
    try:
        return int(float(value))
    except (TypeError, ValueError):
        raise ValueError(
            f"El valor '{value}' de '{campo}' en la fila {fila} no es un número válido"
        )


def leer_pacientes(path: Optional[Path] = None) -> List[Paciente]:
    """Lee el Excel de pacientes y devuelve la lista de :class:`Paciente`.

    Ignora filas totalmente vacías. Si el archivo no existe, levanta
    ``FileNotFoundError`` con un mensaje descriptivo.
    """
    path = path or ruta_excel_pacientes()
    if not path.exists():
        raise FileNotFoundError(
            f"No se encontró el Excel de pacientes en {path}. "
            f"Creá un archivo '{NOMBRE_ARCHIVO_PACIENTES}' en tu escritorio con las columnas: "
            f"nombre y apellido, cuit, numero de sesiones, honorarios por sesion, medio de pago, total"
        )

    logger.info(f"Leyendo Excel de pacientes desde {path}")
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active

    pacientes: List[Paciente] = []
    # min_row=2 para saltear la fila de encabezados
    for fila_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        # Saltear filas completamente vacías
        if row is None or all(celda is None or str(celda).strip() == "" for celda in row):
            continue

        nombre = row[0]
        cuit = row[1]
        numero_de_sesiones = _to_int(row[2], "numero de sesiones", fila_idx)
        honorarios_por_sesion = _to_int(row[3], "honorarios por sesion", fila_idx)

        if not nombre or not str(nombre).strip():
            raise ValueError(f"Falta 'nombre y apellido' en la fila {fila_idx}")

        medio_de_pago = ""
        if len(row) > 4 and row[4] not in (None, ""):
            medio_de_pago = str(row[4]).strip()

        paciente = Paciente(
            nombre_y_apellido=str(nombre).strip(),
            cuit=str(cuit).strip() if cuit not in (None, "") else "",
            numero_de_sesiones=numero_de_sesiones,
            honorarios_por_sesion=honorarios_por_sesion,
            medio_de_pago=medio_de_pago,
        )

        # Validar el total si vino informado en la sexta columna
        if len(row) > 5 and row[5] not in (None, ""):
            total_declarado = _to_int(row[5], "total", fila_idx)
            if total_declarado != paciente.total:
                logger.warning(
                    f"El total declarado ({total_declarado}) para '{paciente.nombre_y_apellido}' "
                    f"no coincide con sesiones * honorarios ({paciente.total}). "
                    f"Se usará el calculado: {paciente.total}"
                )

        pacientes.append(paciente)

    logger.info(f"Se leyeron {len(pacientes)} pacientes del Excel")
    return pacientes
